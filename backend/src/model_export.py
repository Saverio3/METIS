"""
Functions for exporting and importing models to/from Excel.
"""

import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import traceback

def export_model_to_excel(model, file_path=None, include_decomp=True):
    """
    Export model details to an Excel file with multiple sheets.

    Parameters:
    -----------
    model : LinearModel
        The model to export
    file_path : str, optional
        Path where to save the Excel file. If None, uses model name.
    include_decomp : bool, optional
        Whether to include decomposition sheets in the export

    Returns:
    --------
    str
        Path to the saved Excel file
    """
    if model is None or model.results is None:
        print("No valid model to export.")
        return None

    try:
        # Get model coefficients and stats
        params = model.results.params
        tvalues = model.results.tvalues
        pvalues = model.results.pvalues
        conf_int = model.results.conf_int()

        # Create coefficients dataframe
        coef_data = []

        # Add constant term
        if 'const' in params:
            coef_data.append({
                'Variable': 'const',
                'Coefficient': params['const'],
                'T-statistic': tvalues['const'],
                'P-value': pvalues['const'],
                'Lower CI (95%)': conf_int.loc['const'][0],
                'Upper CI (95%)': conf_int.loc['const'][1]
            })

        # Add features
        for feature in model.features:
            if feature in params:
                # Determine if this is an adstock variable and extract the rate
                adstock_rate = 0
                if '_adstock_' in feature:
                    try:
                        adstock_rate = int(feature.split('_adstock_')[1])
                    except:
                        pass

                # Get transformation if available
                transformation = 'None'
                if hasattr(model, 'feature_transformations') and feature in model.feature_transformations:
                    transformation = model.feature_transformations[feature]
                elif hasattr(model, 'loader') and model.loader is not None and hasattr(model.loader, 'get_transformation'):
                    trans = model.loader.get_transformation(feature)
                    if trans:
                        transformation = trans

                coef_data.append({
                    'Variable': feature,
                    'Coefficient': params[feature],
                    'T-statistic': tvalues[feature],
                    'P-value': pvalues[feature],
                    'Lower CI (95%)': conf_int.loc[feature][0],
                    'Upper CI (95%)': conf_int.loc[feature][1],
                    'Adstock Rate': adstock_rate,
                    'Transformation': transformation
                })

        coef_df = pd.DataFrame(coef_data)

        # Create model statistics dataframe
        stats_data = {
            'Statistic': [
                'R-squared',
                'Adjusted R-squared',
                'F-statistic',
                'Prob (F-statistic)',
                'Log-Likelihood',
                'AIC',
                'BIC',
                'No. Observations',
                'Df Residuals',
                'Df Model'
            ],
            'Value': [
                model.results.rsquared,
                model.results.rsquared_adj,
                model.results.fvalue,
                model.results.f_pvalue,
                model.results.llf,
                model.results.aic,
                model.results.bic,
                model.results.nobs,
                model.results.df_resid,
                model.results.df_model
            ]
        }

        # Add Durbin-Watson statistic if available
        try:
            from statsmodels.stats.stattools import durbin_watson
            dw = durbin_watson(model.results.resid)
            stats_data['Statistic'].append('Durbin-Watson')
            stats_data['Value'].append(dw)
        except:
            pass

        # Add Jarque-Bera test if available
        try:
            import statsmodels.api as sm
            jb_stat, jb_pval = sm.stats.jarque_bera(model.results.resid)
            stats_data['Statistic'].append('Jarque-Bera')
            stats_data['Value'].append(jb_stat)
            stats_data['Statistic'].append('Prob (Jarque-Bera)')
            stats_data['Value'].append(jb_pval)
        except:
            pass

        stats_df = pd.DataFrame(stats_data)

        # If file_path is not provided, create one based on model name
        if file_path is None:
            # Create models directory if it doesn't exist
            Path('models').mkdir(parents=True, exist_ok=True)
            file_path = os.path.join('models', f"{model.name}_summary.xlsx")

        # Create a Pandas Excel writer
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Write model metadata
            metadata = pd.DataFrame({
                'Property': ['Model Name', 'KPI', 'Date Range', 'Number of Features', 'Features'],
                'Value': [
                    model.name,
                    model.kpi,
                    f"{model.start_date} to {model.end_date}" if model.start_date and model.end_date else "Full range",
                    len(model.features),
                    ', '.join(model.features)
                ]
            })
            metadata.to_excel(writer, sheet_name='Model Info', index=False)

            # Write coefficients
            coef_df.to_excel(writer, sheet_name='Coefficients', index=False)

            # Write model statistics
            stats_df.to_excel(writer, sheet_name='Model Statistics', index=False)

            # Write model summary in a more detailed form
            # Convert the statsmodels summary to a DataFrame
            summary_text = str(model.get_summary())
            summary_df = pd.DataFrame({'Summary': [line for line in summary_text.split('\n')]})
            summary_df.to_excel(writer, sheet_name='Full Summary', index=False)

            # Add residuals analysis
            residuals = pd.DataFrame({
                'Actual': model.model_data[model.kpi],
                'Predicted': model.results.predict(),
                'Residual': model.results.resid
            })
            residuals.to_excel(writer, sheet_name='Residuals', index=True)

            # Create dataset for variable transformations if any exist
            transform_data = []

            # Add transformations from var_transformations dictionary
            if hasattr(model, 'var_transformations') and model.var_transformations:
                for var_name, info in model.var_transformations.items():
                    if info['type'] == 'split_by_date':
                        start_str = info['start_date'].strftime('%Y-%m-%d') if hasattr(info['start_date'], 'strftime') else str(info['start_date'])
                        end_str = info['end_date'].strftime('%Y-%m-%d') if hasattr(info['end_date'], 'strftime') else str(info['end_date'])

                        transform_data.append({
                            'Variable Name': var_name,
                            'Type': 'split_by_date',
                            'Original Variable': info['original_var'],
                            'Start Date': start_str,
                            'End Date': end_str,
                            'Identifier': info['identifier']
                        })

                    elif info['type'] == 'multiply':
                        transform_data.append({
                            'Variable Name': var_name,
                            'Type': 'multiply',
                            'Variable 1': info['var1'],
                            'Variable 2': info['var2'],
                            'Identifier': info['identifier']
                        })

            # Add adstock transformations
            for feature in model.features:
                if '_adstock_' in feature:
                    try:
                        base_var = feature.split('_adstock_')[0]
                        adstock_rate = int(feature.split('_adstock_')[1])

                        transform_data.append({
                            'Variable Name': feature,
                            'Type': 'adstock',
                            'Original Variable': base_var,
                            'Adstock Rate': adstock_rate
                        })
                    except:
                        pass

            if transform_data:
                transform_df = pd.DataFrame(transform_data)
                transform_df.to_excel(writer, sheet_name='Variable Transformations', index=False)

            # Add a new sheet for all transformed variables - ENHANCED VERSION
            # This will include ALL transformations in the data, not just those used in the model
            if hasattr(model, 'model_data'):
                # Find all transformations in the model data and loader data
                all_columns = list(model.model_data.columns)

                # Get transformed variables by examining column names
                split_vars = [var for var in all_columns if '|SPLIT' in var]
                mult_vars = [var for var in all_columns if '|MULT' in var]
                lag_vars = [var for var in all_columns if '|LAG' in var]
                lead_vars = [var for var in all_columns if '|LEAD' in var]
                adstock_vars = [var for var in all_columns if '_adstock_' in var]
                # Add ICP and ADBUG curve types
                icp_vars = [var for var in all_columns if '|ICP' in var]
                adbug_vars = [var for var in all_columns if '|ADBUG' in var]
                # Add weighted variables
                wgtd_vars = [var for var in all_columns if '|WGTD' in var]

                all_transformed = split_vars + mult_vars + lag_vars + lead_vars + adstock_vars + icp_vars + adbug_vars + wgtd_vars

                # Remove duplicates and sort
                all_transformed = sorted(list(set(all_transformed)))

                if all_transformed:
                    # Create a DataFrame listing all transformed variables
                    transform_list = []

                    for var in all_transformed:
                        # Determine transformation type
                        if '|SPLIT' in var:
                            transform_type = 'split_by_date'
                            parts = var.split('|SPLIT')
                            base_var = parts[0]
                            identifier = parts[1].strip() if len(parts) > 1 else ""
                            transform_list.append({
                                'Variable Name': var,
                                'Transformation Type': transform_type,
                                'Base Variable': base_var,
                                'Identifier': identifier
                            })
                        elif '|MULT' in var:
                            transform_type = 'multiply'
                            parts = var.split('|MULT')
                            base_var = parts[0]
                            identifier = parts[1].strip() if len(parts) > 1 else ""
                            # Further parse the base_var if it contains a multiplication
                            if '*' in base_var:
                                var_parts = base_var.split('*')
                                var1 = var_parts[0].strip()
                                var2 = var_parts[1].strip() if len(var_parts) > 1 else ""
                                transform_list.append({
                                    'Variable Name': var,
                                    'Transformation Type': transform_type,
                                    'Base Variable': base_var,
                                    'Variable 1': var1,
                                    'Variable 2': var2,
                                    'Identifier': identifier
                                })
                            else:
                                transform_list.append({
                                    'Variable Name': var,
                                    'Transformation Type': transform_type,
                                    'Base Variable': base_var,
                                    'Identifier': identifier
                                })
                        elif '|LAG' in var:
                            transform_type = 'lag'
                            parts = var.split('|LAG')
                            base_var = parts[0]
                            period = parts[1].strip() if len(parts) > 1 else "1"
                            transform_list.append({
                                'Variable Name': var,
                                'Transformation Type': transform_type,
                                'Base Variable': base_var,
                                'Period': period
                            })
                        elif '|LEAD' in var:
                            transform_type = 'lead'
                            parts = var.split('|LEAD')
                            base_var = parts[0]
                            period = parts[1].strip() if len(parts) > 1 else "1"
                            transform_list.append({
                                'Variable Name': var,
                                'Transformation Type': transform_type,
                                'Base Variable': base_var,
                                'Period': period
                            })
                        elif '_adstock_' in var:
                            transform_type = 'adstock'
                            parts = var.split('_adstock_')
                            base_var = parts[0]
                            adstock_rate = parts[1] if len(parts) > 1 else "0"
                            transform_list.append({
                                'Variable Name': var,
                                'Transformation Type': transform_type,
                                'Base Variable': base_var,
                                'Adstock Rate': adstock_rate
                            })
                        elif '|ICP' in var:
                            transform_type = 'ICP curve'
                            parts = var.split('|ICP')
                            base_var = parts[0]
                            params = parts[1].strip() if len(parts) > 1 else ""
                            # Extract parameters (a3_b3_g2600 format)
                            alpha, beta, gamma = None, None, None
                            if '_a' in params and '_b' in params and '_g' in params:
                                try:
                                    alpha_part = params.split('_a')[1].split('_b')[0]
                                    beta_part = params.split('_b')[1].split('_g')[0]
                                    gamma_part = params.split('_g')[1]
                                    alpha = alpha_part
                                    beta = beta_part
                                    gamma = gamma_part
                                except:
                                    pass
                            transform_list.append({
                                'Variable Name': var,
                                'Transformation Type': transform_type,
                                'Base Variable': base_var,
                                'Alpha': alpha,
                                'Beta': beta,
                                'Gamma': gamma
                            })
                        elif '|ADBUG' in var:
                            transform_type = 'ADBUG curve'
                            parts = var.split('|ADBUG')
                            base_var = parts[0]
                            params = parts[1].strip() if len(parts) > 1 else ""
                            # Extract parameters (a1_b2_g6000 format)
                            alpha, beta, gamma = None, None, None
                            if '_a' in params and '_b' in params and '_g' in params:
                                try:
                                    alpha_part = params.split('_a')[1].split('_b')[0]
                                    beta_part = params.split('_b')[1].split('_g')[0]
                                    gamma_part = params.split('_g')[1]
                                    alpha = alpha_part
                                    beta = beta_part
                                    gamma = gamma_part
                                except:
                                    pass
                            transform_list.append({
                                'Variable Name': var,
                                'Transformation Type': transform_type,
                                'Base Variable': base_var,
                                'Alpha': alpha,
                                'Beta': beta,
                                'Gamma': gamma
                            })
                        elif '|WGTD' in var:
                            transform_type = 'weighted'
                            parts = var.split('|WGTD')
                            base_var = parts[0]

                            # Get components and coefficients from wgtd_variables if available
                            components = {}
                            if hasattr(model, 'wgtd_variables') and var in model.wgtd_variables:
                                components = model.wgtd_variables[var].get('components', {})

                            transform_list.append({
                                'Variable Name': var,
                                'Transformation Type': transform_type,
                                'Base Variable': base_var,
                                'Components': str(components) if components else ""
                            })

                    # Create and write the DataFrame if we found transformations
                    if transform_list:
                        all_transforms_df = pd.DataFrame(transform_list)
                        all_transforms_df.to_excel(writer, sheet_name='All Transformations', index=False)

            # Export weighted variables separately if they exist
            if hasattr(model, 'wgtd_variables') and model.wgtd_variables:
                weighted_var_data = []
                for var_name, info in model.wgtd_variables.items():
                    base_name = info.get('base_name', '')
                    components = info.get('components', {})

                    # Add a row for each component variable
                    for component, coefficient in components.items():
                        weighted_var_data.append({
                            'Weighted Variable': var_name,
                            'Base Name': base_name,
                            'Component Variable': component,
                            'Coefficient': coefficient
                        })

                if weighted_var_data:
                    weighted_var_df = pd.DataFrame(weighted_var_data)
                    weighted_var_df.to_excel(writer, sheet_name='Weighted Variables', index=False)

        # After the pandas ExcelWriter is closed, we can add more styling using openpyxl directly
        export_transformations_to_excel(model, file_path)

        # Add decomposition sheets if requested
        if include_decomp:
            try:
                from src.model_export import add_decomposition_sheets
                add_decomposition_sheets(model, file_path)
                print("Added decomposition sheets to the Excel file")
            except Exception as e:
                print(f"Warning: Could not add decomposition sheets: {str(e)}")

        print(f"Model exported to Excel: {file_path}")
        return file_path

    except Exception as e:
        print(f"Error exporting model to Excel: {str(e)}")
        traceback.print_exc()  # Print full traceback for debugging
        return None

def import_model_from_excel(loader, filtered_data, excel_path):
    """
    Import model specifications from Excel to recreate a model.

    Parameters:
    -----------
    loader : DataLoader
        The data loader with loaded data
    filtered_data : pandas.DataFrame, optional
        Filtered data if available
    excel_path : str
        Path to the Excel file

    Returns:
    --------
    LinearModel
        The recreated model
    """
    try:
        # Import necessary modules
        from src.linear_models import LinearModel
        from src.model_operations import apply_adstock
        from src.variable_transformations import split_by_date, multiply_variables
        from src.curve_transformations import apply_icp_curve, apply_adbug_curve
        import pandas as pd

        # Try to enhance the LinearModel class
        try:
            from src.compatibility import enhance_model_class
            enhance_model_class()
        except Exception as e:
            print(f"Note: Enhancement not applied: {str(e)}")

        # Read the model info sheet
        model_info = pd.read_excel(excel_path, sheet_name='Model Info', engine='openpyxl')

        # Extract model parameters
        model_name = model_info.loc[model_info['Property'] == 'Model Name', 'Value'].iloc[0]
        kpi_name = model_info.loc[model_info['Property'] == 'KPI', 'Value'].iloc[0]
        features_str = model_info.loc[model_info['Property'] == 'Features', 'Value'].iloc[0]

        # Parse features
        features = [f.strip() for f in features_str.split(',') if f.strip()]

        print(f"Found model '{model_name}' in Excel file")
        print(f"KPI: {kpi_name}")
        print(f"Features: {', '.join(features)}")

        # Use filtered data if available, otherwise use the full dataset
        data_to_use = filtered_data if filtered_data is not None else loader.get_data()

        if data_to_use is None:
            print("No data available. Please load data first.")
            return None

        # Create a new model instance
        model = LinearModel(name=model_name, loader=loader)

        # Initialize attributes if they don't exist
        if not hasattr(model, 'feature_transformations'):
            model.feature_transformations = {}
        if not hasattr(model, 'transformed_data'):
            model.transformed_data = {}
        if not hasattr(model, 'var_transformations'):
            model.var_transformations = {}
        if not hasattr(model, 'wgtd_variables'):
            model.wgtd_variables = {}

        # Set the data first
        model.set_data(data_to_use)

        # Import weighted variables if available
        try:
            wb = pd.ExcelFile(excel_path)
            if 'Weighted Variables' in wb.sheet_names:
                wgtd_df = pd.read_excel(excel_path, sheet_name='Weighted Variables')

                # Process weighted variables
                wgtd_vars = {}
                for _, row in wgtd_df.iterrows():
                    wgt_var = row['Weighted Variable']
                    base_name = row['Base Name']
                    comp_var = row['Component Variable']
                    coef = row['Coefficient']

                    if wgt_var not in wgtd_vars:
                        wgtd_vars[wgt_var] = {'base_name': base_name, 'components': {}}

                    wgtd_vars[wgt_var]['components'][comp_var] = coef

                # Add to model
                model.wgtd_variables = wgtd_vars

                # Create weighted variables
                for var_name, info in wgtd_vars.items():
                    base_name = info['base_name']
                    components = info['components']

                    # Initialize the weighted variable
                    model.model_data[var_name] = 0.0

                    # Add each component
                    for comp_var, coef in components.items():
                        if comp_var in model.model_data.columns:
                            model.model_data[var_name] += float(coef) * model.model_data[comp_var]
                        else:
                            print(f"Warning: Component variable '{comp_var}' for weighted variable '{var_name}' not found.")

                    # Register with loader
                    if loader is not None:
                        # Check if loader has data or _data attribute
                        if hasattr(loader, 'data'):
                            loader.data[var_name] = model.model_data[var_name]
                        elif hasattr(loader, '_data'):
                            loader._data[var_name] = model.model_data[var_name]
                        else:
                            print(f"Warning: Cannot update loader with weighted variable '{var_name}', no data attribute found")

                        # Create a transformation registry if needed
                        if not hasattr(loader, 'transformations_registry'):
                            loader.transformations_registry = {}

                        # Add weighted variable to the registry
                        loader.transformations_registry[var_name] = {
                            'type': 'weighted',
                            'base_name': base_name,
                            'components': components,
                            'is_transformed': True
                        }

                    print(f"Created weighted variable: {var_name}")
        except Exception as e:
            print(f"Warning: Error loading weighted variables: {str(e)}")

        # Process features to identify transformations
        for feature in features:
            # Handle adstock transformations
            if '_adstock_' in feature:
                try:
                    base_var = feature.split('_adstock_')[0]
                    adstock_rate = int(feature.split('_adstock_')[1])/100

                    # Apply adstock transformation
                    if base_var in model.model_data.columns:
                        model.model_data[feature] = apply_adstock(
                            model.model_data[base_var],
                            adstock_rate
                        )
                        # Record transformation
                        model.var_transformations[feature] = {
                            'type': 'adstock',
                            'original_var': base_var,
                            'adstock_rate': adstock_rate
                        }

                        # Register with loader
                        if loader is not None:
                            # Make sure data is updated in loader - fixed to check for both data and _data attributes
                            if hasattr(loader, 'data'):
                                loader.data[feature] = model.model_data[feature]
                            elif hasattr(loader, '_data'):
                                loader._data[feature] = model.model_data[feature]
                            else:
                                print(f"Warning: Cannot update loader with adstock variable '{feature}', no data attribute found")

                            # Create a transformation registry if needed
                            if not hasattr(loader, 'transformations_registry'):
                                loader.transformations_registry = {}

                            # Add to the registry
                            loader.transformations_registry[feature] = {
                                'type': 'adstock',
                                'base_variable': base_var,
                                'adstock_rate': adstock_rate * 100,  # Convert back to percentage
                                'is_transformed': True
                            }

                        print(f"Applied adstock transformation: {feature}")
                    else:
                        print(f"Warning: Base variable '{base_var}' for adstock not found")
                except Exception as e:
                    print(f"Error applying adstock to {feature}: {str(e)}")

        # Process variables from All Transformations sheet
        try:
            # Load All Transformations sheet
            transform_df = pd.read_excel(excel_path, sheet_name='All Transformations', engine='openpyxl')

            if not transform_df.empty:
                print(f"Found {len(transform_df)} transformations in 'All Transformations' sheet")

                for _, row in transform_df.iterrows():
                    var_name = row['Variable Name']
                    transform_type = row['Transformation Type']

                    # Skip if the variable already exists
                    if var_name in model.model_data.columns:
                        continue

                    # Apply transformation based on type
                    if transform_type == 'lag' and 'Period' in row and pd.notnull(row['Period']):
                        base_var = row['Base Variable']
                        period = int(row['Period'])
                        if base_var in model.model_data.columns:
                            model.model_data[var_name] = model.model_data[base_var].shift(period)
                            model.var_transformations[var_name] = {
                                'type': 'lag',
                                'original_var': base_var,
                                'period': period
                            }

                            # Register with loader
                            if loader is not None:
                                # Update loader data - fixed to check for both data and _data attributes
                                if hasattr(loader, 'data'):
                                    loader.data[var_name] = model.model_data[var_name]
                                elif hasattr(loader, '_data'):
                                    loader._data[var_name] = model.model_data[var_name]
                                else:
                                    print(f"Warning: Cannot update loader with variable '{var_name}', no data attribute found")

                                # Create a transformation registry if needed
                                if not hasattr(loader, 'transformations_registry'):
                                    loader.transformations_registry = {}

                                # Add to registry
                                loader.transformations_registry[var_name] = {
                                    'type': 'lag',
                                    'base_variable': base_var,
                                    'period': period,
                                    'is_transformed': True
                                }

                            print(f"Applied lag transformation from All Transformations: {var_name}")

                    elif transform_type == 'lead' and 'Period' in row and pd.notnull(row['Period']):
                        base_var = row['Base Variable']
                        period = int(row['Period'])
                        if base_var in model.model_data.columns:
                            model.model_data[var_name] = model.model_data[base_var].shift(-period)
                            model.var_transformations[var_name] = {
                                'type': 'lead',
                                'original_var': base_var,
                                'period': period
                            }

                            # Register with loader
                            if loader is not None:
                                # Update loader data - fixed to check for both data and _data attributes
                                if hasattr(loader, 'data'):
                                    loader.data[var_name] = model.model_data[var_name]
                                elif hasattr(loader, '_data'):
                                    loader._data[var_name] = model.model_data[var_name]
                                else:
                                    print(f"Warning: Cannot update loader with variable '{var_name}', no data attribute found")

                                # Create a transformation registry if needed
                                if not hasattr(loader, 'transformations_registry'):
                                    loader.transformations_registry = {}

                                # Add to registry
                                loader.transformations_registry[var_name] = {
                                    'type': 'lead',
                                    'base_variable': base_var,
                                    'period': period,
                                    'is_transformed': True
                                }

                            print(f"Applied lead transformation from All Transformations: {var_name}")

                    elif transform_type == 'split_by_date':
                        base_var = row['Base Variable']
                        identifier = row['Identifier'] if 'Identifier' in row and pd.notnull(row['Identifier']) else ""
                        start_date = row['Start Date'] if 'Start Date' in row and pd.notnull(row['Start Date']) else None
                        end_date = row['End Date'] if 'End Date' in row and pd.notnull(row['End Date']) else None

                        if start_date:
                            start_date = pd.to_datetime(start_date)
                        if end_date:
                            end_date = pd.to_datetime(end_date)

                        if base_var in model.model_data.columns:
                            _, _ = split_by_date(
                                model.model_data,
                                base_var,
                                start_date,
                                end_date,
                                identifier,
                                inplace=True
                            )
                            model.var_transformations[var_name] = {
                                'type': 'split_by_date',
                                'original_var': base_var,
                                'start_date': start_date,
                                'end_date': end_date,
                                'identifier': identifier
                            }

                            # Register with loader
                            if loader is not None:
                                # Update loader data - fixed to check for both data and _data attributes
                                if hasattr(loader, 'data'):
                                    loader.data[var_name] = model.model_data[var_name]
                                elif hasattr(loader, '_data'):
                                    loader._data[var_name] = model.model_data[var_name]
                                else:
                                    print(f"Warning: Cannot update loader with variable '{var_name}', no data attribute found")

                                # Create a transformation registry if needed
                                if not hasattr(loader, 'transformations_registry'):
                                    loader.transformations_registry = {}

                                # Add to registry
                                loader.transformations_registry[var_name] = {
                                    'type': 'split_by_date',
                                    'base_variable': base_var,
                                    'start_date': start_date,
                                    'end_date': end_date,
                                    'identifier': identifier,
                                    'is_transformed': True
                                }

                            print(f"Applied split transformation from All Transformations: {var_name}")

                    elif transform_type == 'multiply':
                        if 'Variable 1' in row and 'Variable 2' in row:
                            var1 = row['Variable 1']
                            var2 = row['Variable 2']
                            identifier = row['Identifier'] if 'Identifier' in row and pd.notnull(row['Identifier']) else ""

                            if var1 in model.model_data.columns and var2 in model.model_data.columns:
                                _, _ = multiply_variables(
                                    model.model_data,
                                    var1,
                                    var2,
                                    identifier,
                                    inplace=True
                                )
                                model.var_transformations[var_name] = {
                                    'type': 'multiply',
                                    'var1': var1,
                                    'var2': var2,
                                    'identifier': identifier
                                }

                                # Register with loader
                                if loader is not None:
                                    # Update loader data - fixed to check for both data and _data attributes
                                    if hasattr(loader, 'data'):
                                        loader.data[var_name] = model.model_data[var_name]
                                    elif hasattr(loader, '_data'):
                                        loader._data[var_name] = model.model_data[var_name]
                                    else:
                                        print(f"Warning: Cannot update loader with variable '{var_name}', no data attribute found")

                                    # Create a transformation registry if needed
                                    if not hasattr(loader, 'transformations_registry'):
                                        loader.transformations_registry = {}

                                    # Add to registry
                                    loader.transformations_registry[var_name] = {
                                        'type': 'multiply',
                                        'var1': var1,
                                        'var2': var2,
                                        'identifier': identifier,
                                        'is_transformed': True
                                    }

                                print(f"Applied multiply transformation from All Transformations: {var_name}")

                    elif transform_type == 'ICP curve' or transform_type == 'ICP':
                        base_var = row['Base Variable']
                        alpha = float(row['Alpha']) if 'Alpha' in row and pd.notnull(row['Alpha']) else 3.0
                        beta = float(row['Beta']) if 'Beta' in row and pd.notnull(row['Beta']) else 3.0
                        gamma = float(row['Gamma']) if 'Gamma' in row and pd.notnull(row['Gamma']) else 100.0

                        if base_var in model.model_data.columns:
                            model.model_data[var_name] = apply_icp_curve(
                                model.model_data[base_var],
                                alpha,
                                beta,
                                gamma
                            )
                            model.var_transformations[var_name] = {
                                'type': 'ICP',
                                'original_var': base_var,
                                'alpha': alpha,
                                'beta': beta,
                                'gamma': gamma
                            }

                            # Register with loader
                            if loader is not None:
                                # Update loader data - fixed to check for both data and _data attributes
                                if hasattr(loader, 'data'):
                                    loader.data[var_name] = model.model_data[var_name]
                                elif hasattr(loader, '_data'):
                                    loader._data[var_name] = model.model_data[var_name]
                                else:
                                    print(f"Warning: Cannot update loader with variable '{var_name}', no data attribute found")

                                # Create a transformation registry if needed
                                if not hasattr(loader, 'transformations_registry'):
                                    loader.transformations_registry = {}

                                # Add to registry
                                loader.transformations_registry[var_name] = {
                                    'type': 'ICP',
                                    'base_variable': base_var,
                                    'alpha': alpha,
                                    'beta': beta,
                                    'gamma': gamma,
                                    'is_transformed': True
                                }

                            print(f"Applied ICP curve transformation from All Transformations: {var_name}")

                    elif transform_type == 'ADBUG curve' or transform_type == 'ADBUG':
                        base_var = row['Base Variable']
                        alpha = float(row['Alpha']) if 'Alpha' in row and pd.notnull(row['Alpha']) else 1.0
                        beta = float(row['Beta']) if 'Beta' in row and pd.notnull(row['Beta']) else 2.0
                        gamma = float(row['Gamma']) if 'Gamma' in row and pd.notnull(row['Gamma']) else 100.0

                        if base_var in model.model_data.columns:
                            model.model_data[var_name] = apply_adbug_curve(
                                model.model_data[base_var],
                                alpha,
                                beta,
                                gamma
                            )
                            model.var_transformations[var_name] = {
                                'type': 'ADBUG',
                                'original_var': base_var,
                                'alpha': alpha,
                                'beta': beta,
                                'gamma': gamma
                            }

                            # Register with loader
                            if loader is not None:
                                # Update loader data - fixed to check for both data and _data attributes
                                if hasattr(loader, 'data'):
                                    loader.data[var_name] = model.model_data[var_name]
                                elif hasattr(loader, '_data'):
                                    loader._data[var_name] = model.model_data[var_name]
                                else:
                                    print(f"Warning: Cannot update loader with variable '{var_name}', no data attribute found")

                                # Create a transformation registry if needed
                                if not hasattr(loader, 'transformations_registry'):
                                    loader.transformations_registry = {}

                                # Add to registry
                                loader.transformations_registry[var_name] = {
                                    'type': 'ADBUG',
                                    'base_variable': base_var,
                                    'alpha': alpha,
                                    'beta': beta,
                                    'gamma': gamma,
                                    'is_transformed': True
                                }

                            print(f"Applied ADBUG curve transformation from All Transformations: {var_name}")

        except Exception as e:
            print(f"Note: Could not load all transformations from Excel: {str(e)}")
            import traceback
            traceback.print_exc()

        # CRITICAL: Update the loader's data with all the transformations
        if loader is not None:
            # Update the loader's data with all variables from the model
            for col in model.model_data.columns:
                # Fixed to check for both data and _data attributes
                if hasattr(loader, 'data'):
                    loader.data[col] = model.model_data[col]
                elif hasattr(loader, '_data'):
                    loader._data[col] = model.model_data[col]
                else:
                    print(f"Warning: Cannot update loader with variable '{col}', no data attribute found")

                # If this is a transformed variable, make sure it's marked as such
                if col not in data_to_use.columns or '|' in col or '_adstock_' in col:
                    # Determine if a base variable exists
                    base_var = None
                    if '|' in col:
                        base_var = col.split('|')[0]
                    elif '_adstock_' in col:
                        base_var = col.split('_adstock_')[0]

                    # Ensure transformation registry exists
                    if not hasattr(loader, 'transformations_registry'):
                        loader.transformations_registry = {}

                    # Add basic transformation info if not already added
                    if col not in loader.transformations_registry:
                        loader.transformations_registry[col] = {
                            'type': 'unknown',
                            'base_variable': base_var,
                            'is_transformed': True
                        }

        # Set the KPI - Case-insensitive handling
        kpi_found = False
        for col in model.model_data.columns:
            if col.lower() == kpi_name.lower():
                model.kpi = col
                kpi_found = True
                print(f"Setting KPI to: {col}")
                break

        if not kpi_found:
            print(f"Error: KPI '{kpi_name}' not found in the data")
            # Still set the KPI to maintain model structure
            model.kpi = kpi_name

        # Clean features list - remove any features not in the data
        cleaned_features = []
        for feature in features:
            # Case-insensitive feature checking
            feature_found = False
            for col in model.model_data.columns:
                if col.lower() == feature.lower():
                    # Use the actual column name from the data
                    cleaned_features.append(col)
                    feature_found = True
                    break

            if not feature_found:
                print(f"Warning: Feature '{feature}' not found in model data, removing from features list")

        model.features = cleaned_features
        print(f"Final features list: {model.features}")

        # Only try to fit the model if we have KPI and at least some features
        if model.kpi in model.model_data.columns and model.features:
            # Fit the model with all features
            try:
                import statsmodels.api as sm
                # Prepare the data
                y = model.model_data[model.kpi]

                # Prepare X data with transformations if available
                X_data = {}
                for feature in model.features:
                    if hasattr(model, 'transformed_data') and feature in model.transformed_data:
                        X_data[feature] = model.transformed_data[feature]
                    else:
                        X_data[feature] = model.model_data[feature]

                # Create DataFrame and add constant
                X = pd.DataFrame(X_data, index=y.index)
                X = sm.add_constant(X)

                # Check for nulls
                has_nulls = X.isnull().any().any() or y.isnull().any()
                if has_nulls:
                    print("Warning: Data contains null values. Dropping rows with nulls.")
                    # Get common index of non-null values
                    non_null_idx = y.dropna().index.intersection(X.dropna().index)
                    X = X.loc[non_null_idx]
                    y = y.loc[non_null_idx]

                # Debug info
                print(f"X shape: {X.shape}, y shape: {y.shape}")
                if X.empty or y.empty:
                    print("ERROR: No data available after filtering nulls")
                    return model

                # Fit the model
                model.model = sm.OLS(y, X)
                model.results = model.model.fit()

                print(f"\nModel recreated successfully with {len(model.features)} features!")
                print(f"R-squared: {model.results.rsquared:.4f}")
            except Exception as e:
                print(f"Error fitting model: {str(e)}")
                import traceback
                traceback.print_exc()
        else:
            print("Not enough data to fit the model - missing KPI or features")

        return model

    except Exception as e:
        print(f"Error importing model from Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

"""
Functions to extend Excel export/import with variable transformation information.
"""

def export_transformations_to_excel(model, excel_path):
    """
    Add variable transformations to an Excel export file.

    Parameters:
    -----------
    model : LinearModel
        Model containing the transformations
    excel_path : str
        Path to the Excel file

    Returns:
    --------
    bool
        True if successful, False otherwise
    """
    if not hasattr(model, 'var_transformations') or not model.var_transformations:
        print("No variable transformations to export.")
        return True  # Return True even if nothing to export

    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_path)

        # Create a new sheet for transformations
        if 'Variable Transformations' in wb.sheetnames:
            # Remove existing sheet
            del wb['Variable Transformations']

        # Create new sheet
        ws = wb.create_sheet(title='Variable Transformations')

        # Set up headers with styling
        headers = ['Variable Name', 'Transformation Type', 'Parameters']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Add transformations
        row = 2
        for var_name, info in model.var_transformations.items():
            # Variable name
            ws.cell(row=row, column=1).value = var_name

            # Transformation type
            ws.cell(row=row, column=2).value = info['type']

            # Parameters as JSON string
            if info['type'] == 'split_by_date':
                # For split_by_date transformations
                start_str = info['start_date'].strftime('%Y-%m-%d') if hasattr(info['start_date'], 'strftime') else str(info['start_date'])
                end_str = info['end_date'].strftime('%Y-%m-%d') if hasattr(info['end_date'], 'strftime') else str(info['end_date'])

                params = (
                    f"Original Variable: {info['original_var']}, "
                    f"Start Date: {start_str}, "
                    f"End Date: {end_str}, "
                    f"Identifier: {info['identifier']}"
                )

            elif info['type'] == 'multiply':
                # For multiply transformations
                params = (
                    f"Variable 1: {info['var1']}, "
                    f"Variable 2: {info['var2']}, "
                    f"Identifier: {info['identifier']}"
                )
            else:
                params = str(info)

            ws.cell(row=row, column=3).value = params

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50

        # Apply alternating row colors
        for row in range(2, ws.max_row + 1):
            if row % 2 == 0:
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
                    )

        # Save the workbook
        wb.save(excel_path)
        print(f"Variable transformations exported to Excel: {excel_path}")
        return True

    except Exception as e:
        print(f"Error exporting transformations to Excel: {str(e)}")
        return False

def import_transformations_from_excel(model, excel_path):
    """
    Import variable transformations from an Excel file.

    Parameters:
    -----------
    model : LinearModel
        Model to add the transformations to
    excel_path : str
        Path to the Excel file

    Returns:
    --------
    bool
        True if successful, False otherwise
    """
    try:
        # Check if the Excel file exists
        if not os.path.exists(excel_path):
            print(f"Excel file not found: {excel_path}")
            return False

        # Try to read the transformations sheet using pandas first, which is more reliable
        try:
            transform_df = pd.read_excel(excel_path, sheet_name='Variable Transformations', engine='openpyxl')

            # Prepare transformation dictionary
            if not hasattr(model, 'var_transformations'):
                model.var_transformations = {}

            # Import transformations from DataFrame
            for idx, row in transform_df.iterrows():
                var_name = row['Variable Name']
                trans_type = row['Type']

                if trans_type == 'split_by_date':
                    # Extract parameters
                    original_var = row['Original Variable']
                    start_date = row.get('Start Date', None)
                    if start_date is not None and pd.notna(start_date):
                        start_date = pd.to_datetime(start_date)
                    end_date = row.get('End Date', None)
                    if end_date is not None and pd.notna(end_date):
                        end_date = pd.to_datetime(end_date)
                    identifier = row.get('Identifier', '')

                    # Add to transformations
                    model.var_transformations[var_name] = {
                        'type': 'split_by_date',
                        'original_var': original_var,
                        'start_date': start_date,
                        'end_date': end_date,
                        'identifier': identifier
                    }

                elif trans_type == 'multiply':
                    # Extract parameters
                    var1 = row.get('Variable 1', '')
                    var2 = row.get('Variable 2', '')
                    identifier = row.get('Identifier', '')

                    # Add to transformations
                    model.var_transformations[var_name] = {
                        'type': 'multiply',
                        'var1': var1,
                        'var2': var2,
                        'identifier': identifier
                    }

                elif trans_type == 'adstock':
                    # For adstock transformations, we'll handle them in import_model_from_excel
                    # Just record them here for reference
                    original_var = row.get('Original Variable', '')
                    adstock_rate = row.get('Adstock Rate', 0)

                    # Store adstock information
                    model.var_transformations[var_name] = {
                        'type': 'adstock',
                        'original_var': original_var,
                        'adstock_rate': adstock_rate
                    }

            print(f"Imported {len(model.var_transformations)} variable transformations from Excel")
            return True

        except Exception as e:
            # If pandas fails, fall back to openpyxl direct read
            print(f"Warning: Failed to read transformations with pandas, falling back to openpyxl: {str(e)}")

            # Load the workbook with openpyxl
            wb = openpyxl.load_workbook(excel_path, read_only=True)

            # Check if the transformations sheet exists
            if 'Variable Transformations' not in wb.sheetnames:
                print("No variable transformations found in Excel file.")
                return True  # Return True even if nothing to import

            # Access the transformations sheet
            ws = wb['Variable Transformations']

            # Prepare transformation dictionary
            if not hasattr(model, 'var_transformations'):
                model.var_transformations = {}

            # Find the header row and column indexes
            headers = {}
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers[header] = col

            # Check for required headers
            if 'Variable Name' not in headers or 'Type' not in headers:
                print("Missing required headers in Variable Transformations sheet.")
                return False

            # Import transformations
            for row in range(2, ws.max_row + 1):
                var_name = ws.cell(row=row, column=headers['Variable Name']).value
                if not var_name:
                    continue

                trans_type = ws.cell(row=row, column=headers['Type']).value

                if trans_type == 'split_by_date':
                    # Extract parameters
                    orig_var_col = headers.get('Original Variable')
                    orig_var = ws.cell(row=row, column=orig_var_col).value if orig_var_col else None

                    start_date_col = headers.get('Start Date')
                    start_date_val = ws.cell(row=row, column=start_date_col).value if start_date_col else None
                    start_date = None if not start_date_val else pd.to_datetime(start_date_val)

                    end_date_col = headers.get('End Date')
                    end_date_val = ws.cell(row=row, column=end_date_col).value if end_date_col else None
                    end_date = None if not end_date_val else pd.to_datetime(end_date_val)

                    id_col = headers.get('Identifier')
                    identifier = ws.cell(row=row, column=id_col).value if id_col else ""

                    # Add to transformations
                    model.var_transformations[var_name] = {
                        'type': 'split_by_date',
                        'original_var': orig_var,
                        'start_date': start_date,
                        'end_date': end_date,
                        'identifier': identifier or ""
                    }

                elif trans_type == 'multiply':
                    # Extract parameters
                    var1_col = headers.get('Variable 1')
                    var1 = ws.cell(row=row, column=var1_col).value if var1_col else None

                    var2_col = headers.get('Variable 2')
                    var2 = ws.cell(row=row, column=var2_col).value if var2_col else None

                    id_col = headers.get('Identifier')
                    identifier = ws.cell(row=row, column=id_col).value if id_col else ""

                    # Add to transformations
                    model.var_transformations[var_name] = {
                        'type': 'multiply',
                        'var1': var1,
                        'var2': var2,
                        'identifier': identifier or ""
                    }

                elif trans_type == 'adstock':
                    # Extract parameters
                    orig_var_col = headers.get('Original Variable')
                    orig_var = ws.cell(row=row, column=orig_var_col).value if orig_var_col else None

                    rate_col = headers.get('Adstock Rate')
                    adstock_rate = ws.cell(row=row, column=rate_col).value if rate_col else 0

                    # Add to transformations
                    model.var_transformations[var_name] = {
                        'type': 'adstock',
                        'original_var': orig_var,
                        'adstock_rate': adstock_rate
                    }

            print(f"Imported {len(model.var_transformations)} variable transformations from Excel")
            return True

    except Exception as e:
        print(f"Error importing transformations from Excel: {str(e)}")
        traceback.print_exc()
        return False



### Functions to add decomposition sheets

def add_decomposition_sheets(model, excel_path):
    """
    Add decomposition sheets to an existing Excel export file.

    Parameters:
    -----------
    model : LinearModel
        The model to decompose
    excel_path : str
        Path to the Excel file

    Returns:
    --------
    bool
        True if successful, False otherwise
    """
    try:
        import pandas as pd
        import numpy as np
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Check if file exists
        if not os.path.exists(excel_path):
            print(f"Excel file not found: {excel_path}")
            return False

        # First calculate the decompositions
        from src.decomposition import get_variable_groups, create_default_groups, calculate_decomposition

        # Get variable groups
        groups = get_variable_groups(model)

        if not groups:
            groups = create_default_groups(model)
            print("Created default groups for decomposition.")

        # Calculate group-level decomposition
        group_decomp_df = calculate_decomposition(model, groups)

        # Calculate individual variable decomposition
        var_decomp_df = calculate_variable_decomposition(model)

        # Load the workbook
        wb = openpyxl.load_workbook(excel_path)

        # Add group decomposition sheet
        add_decomp_sheet(wb, group_decomp_df, "Group Decomposition")

        # Add variable decomposition sheet
        add_variable_decomp_sheet(wb, var_decomp_df, model, groups, "Variable Decomposition")

        # Save the workbook
        wb.save(excel_path)

        print(f"Decomposition sheets added to Excel: {excel_path}")
        return True

    except Exception as e:
        print(f"Error adding decomposition sheets: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def calculate_variable_decomposition(model):
    """
    Calculate decomposition at the individual variable level.

    Parameters:
    -----------
    model : LinearModel
        The model to decompose

    Returns:
    --------
    pandas.DataFrame
        DataFrame with decomposed contributions for each variable
    """
    # Get model coefficients
    coefficients = model.results.params

    # Get model data
    data = model.model_data.copy()

    # Create DataFrame to store contributions
    contributions = pd.DataFrame(index=data.index)

    # Add actual KPI values
    contributions['Actual'] = data[model.kpi]

    # Add predicted values
    contributions['Predicted'] = model.results.predict()

    # Calculate residuals
    contributions['Residual'] = contributions['Actual'] - contributions['Predicted']

    # Calculate individual variable contributions
    for var in model.features + ['const']:
        # Skip variables not in coefficients
        if var not in coefficients:
            continue

        # Get coefficient
        coef = coefficients[var]

        # Get variable values (handle 'const' specially)
        if var == 'const':
            values = pd.Series(1, index=data.index)
        else:
            # Skip if variable not in data
            if var not in data.columns:
                print(f"Warning: Variable '{var}' not found in data, skipping")
                continue
            values = data[var]

        # Calculate contribution
        contributions[var] = coef * values

    return contributions

def add_decomp_sheet(wb, decomp_df, sheet_name):
    """
    Add a decomposition sheet to the workbook.

    Parameters:
    -----------
    wb : openpyxl.Workbook
        Workbook to add sheet to
    decomp_df : pandas.DataFrame
        DataFrame with decomposition data
    sheet_name : str
        Name for the sheet
    """
    # Create a new sheet for decomposition (or overwrite if exists)
    if sheet_name in wb.sheetnames:
        # Remove existing sheet
        del wb[sheet_name]

    # Create new sheet
    ws = wb.create_sheet(title=sheet_name)

    # Convert date index to string if needed
    if isinstance(decomp_df.index, pd.DatetimeIndex):
        index_values = [date.strftime('%Y-%m-%d') for date in decomp_df.index]
    else:
        index_values = [str(idx) for idx in decomp_df.index]

    # Add header row with column names
    ws.cell(row=1, column=1, value="Observation")
    for col_idx, col_name in enumerate(decomp_df.columns, start=2):
        ws.cell(row=1, column=col_idx, value=col_name)
        # Style header cells
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Add data rows
    for row_idx, (idx, row) in enumerate(decomp_df.iterrows(), start=2):
        # Add index value
        ws.cell(row=row_idx, column=1, value=index_values[row_idx-2])

        # Add values for each column
        for col_idx, (col_name, value) in enumerate(row.items(), start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

def add_variable_decomp_sheet(wb, var_decomp_df, model, groups, sheet_name):
    """
    Add a variable decomposition sheet to the workbook with group headers.

    Parameters:
    -----------
    wb : openpyxl.Workbook
        Workbook to add sheet to
    var_decomp_df : pandas.DataFrame
        DataFrame with variable decomposition data
    model : LinearModel
        Model with variable information
    groups : dict
        Dictionary with group assignments
    sheet_name : str
        Name for the sheet
    """
    # Create a new sheet for variable decomposition (or overwrite if exists)
    if sheet_name in wb.sheetnames:
        # Remove existing sheet
        del wb[sheet_name]

    # Create new sheet
    ws = wb.create_sheet(title=sheet_name)

    # Convert date index to string if needed
    if isinstance(var_decomp_df.index, pd.DatetimeIndex):
        index_values = [date.strftime('%Y-%m-%d') for date in var_decomp_df.index]
    else:
        index_values = [str(idx) for idx in var_decomp_df.index]

    # Group variables by their group assignment
    grouped_vars = {}

    # Expand weighted variables first
    expanded_vars = []
    for var in model.features + ['const']:
        if var in var_decomp_df.columns:
            expanded_vars.append(var)

    # Add components from weighted variables
    if hasattr(model, 'wgtd_variables'):
        for wgtd_var, wgtd_info in model.wgtd_variables.items():
            if wgtd_var in model.features and wgtd_var in var_decomp_df.columns:
                # Get component variables and coefficients
                components = wgtd_info.get('components', {})

                # Add components to expanded vars
                for component in components:
                    if component not in expanded_vars:
                        expanded_vars.append(component)

                # Remove weighted variable
                if wgtd_var in expanded_vars:
                    expanded_vars.remove(wgtd_var)

    # Now group variables
    for var in expanded_vars:
        # Get group for this variable
        group = groups.get(var, {}).get('Group', 'Other')
        if group not in grouped_vars:
            grouped_vars[group] = []
        grouped_vars[group].append(var)

    # Add groups header row
    ws.cell(row=1, column=1, value="Groups")
    current_col = 2

    # Add default columns first
    for col_name in ['Actual', 'Predicted', 'Residual']:
        ws.cell(row=1, column=current_col, value="")
        current_col += 1

    # Add group headers for each variable
    for group, vars_in_group in grouped_vars.items():
        for var in vars_in_group:
            ws.cell(row=1, column=current_col, value=group)
            # Style group header cells
            cell = ws.cell(row=1, column=current_col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

            # Set background color based on group
            if group == 'Base':
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            elif group == 'Media':
                cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            elif group == 'Price':
                cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
            elif group == 'Seasonality':
                cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

            current_col += 1

    # Add variable names row
    ws.cell(row=2, column=1, value="Observation")
    current_col = 2

    # Add default columns
    for col_name in ['Actual', 'Predicted', 'Residual']:
        ws.cell(row=2, column=current_col, value=col_name)
        # Style variable header cells
        cell = ws.cell(row=2, column=current_col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        current_col += 1

    # Add variable headers
    for group, vars_in_group in grouped_vars.items():
        for var in vars_in_group:
            ws.cell(row=2, column=current_col, value=var)
            # Style variable header cells
            cell = ws.cell(row=2, column=current_col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            current_col += 1

    # Add data rows
    for row_idx, (idx, row) in enumerate(var_decomp_df.iterrows(), start=3):
        # Add index value
        ws.cell(row=row_idx, column=1, value=index_values[row_idx-3])

        # Current column position
        current_col = 2

        # Add default columns first
        for col_name in ['Actual', 'Predicted', 'Residual']:
            ws.cell(row=row_idx, column=current_col, value=row[col_name])
            current_col += 1

        # Add variable values
        for group, vars_in_group in grouped_vars.items():
            for var in vars_in_group:
                # If this is a component of a weighted variable, calculate its value
                if hasattr(model, 'wgtd_variables') and var not in var_decomp_df.columns:
                    # Find if this is a component in any weighted variable
                    value = 0
                    for wgtd_var, wgtd_info in model.wgtd_variables.items():
                        if wgtd_var in var_decomp_df.columns and var in wgtd_info.get('components', {}):
                            # Get the weighted variable's contribution
                            wgtd_contrib = var_decomp_df[wgtd_var].iloc[row_idx-3]
                            components = wgtd_info.get('components', {})

                            # Calculate total weight
                            total_weight = sum(abs(coef) for coef in components.values())

                            if total_weight > 0:
                                # Calculate component's proportional contribution
                                component_coef = components[var]
                                component_contrib = wgtd_contrib * (component_coef / total_weight)
                                value += component_contrib

                    ws.cell(row=row_idx, column=current_col, value=value)
                else:
                    # Regular variable
                    if var in var_decomp_df.columns:
                        ws.cell(row=row_idx, column=current_col, value=row[var])
                    else:
                        ws.cell(row=row_idx, column=current_col, value=0)

                current_col += 1

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Freeze the first two rows and first column
    ws.freeze_panes = ws.cell(3, 2)



def export_weighted_variables(model, file_path):
    """
    Export weighted variables to an Excel file.

    Parameters:
    -----------
    model : LinearModel
        The model containing weighted variables
    file_path : str
        Path to the Excel file

    Returns:
    --------
    bool
        True if successful, False otherwise
    """
    if not hasattr(model, 'wgtd_variables') or not model.wgtd_variables:
        return True  # Return True even if no weighted variables to export

    try:
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Load the workbook
        wb = openpyxl.load_workbook(file_path)

        # Check if sheet already exists
        sheet_name = 'Weighted Variables'
        if sheet_name in wb.sheetnames:
            # Remove existing sheet
            del wb[sheet_name]

        # Create a new sheet
        ws = wb.create_sheet(title=sheet_name)

        # Create column headers
        headers = ['Weighted Variable', 'Base Name', 'Component Variable', 'Coefficient']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Add data for each weighted variable
        row_idx = 2
        for wgtd_var, info in model.wgtd_variables.items():
            base_name = info.get('base_name', '')
            components = info.get('components', {})

            # If no components, add just one row with the weighted variable
            if not components:
                ws.cell(row=row_idx, column=1).value = wgtd_var
                ws.cell(row=row_idx, column=2).value = base_name
                row_idx += 1
                continue

            # Add a row for each component
            for component, coefficient in components.items():
                # Always add the weighted variable name and base name
                ws.cell(row=row_idx, column=1).value = wgtd_var
                ws.cell(row=row_idx, column=2).value = base_name

                ws.cell(row=row_idx, column=3).value = component
                ws.cell(row=row_idx, column=4).value = coefficient

                # Apply alternating row colors
                if row_idx % 2 == 0:
                    for col in range(1, 5):
                        ws.cell(row=row_idx, column=col).fill = PatternFill(
                            start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
                        )

                row_idx += 1

        # Set column widths
        ws.column_dimensions['A'].width = 30  # Weighted Variable
        ws.column_dimensions['B'].width = 20  # Base Name
        ws.column_dimensions['C'].width = 30  # Component Variable
        ws.column_dimensions['D'].width = 15  # Coefficient

        # Save the workbook
        wb.save(file_path)
        return True
    except Exception as e:
        print(f"Error exporting weighted variables: {str(e)}")
        return False

def import_weighted_variables(model, excel_path):
    """
    Import weighted variables from an Excel file.

    Parameters:
    -----------
    model : LinearModel
        The model to add weighted variables to
    excel_path : str
        Path to the Excel file

    Returns:
    --------
    bool
        True if successful, False otherwise
    """
    try:
        import pandas as pd
        import os

        # Check if file exists
        if not os.path.exists(excel_path):
            return False

        # Try to read the weighted variables sheet
        try:
            wgtd_df = pd.read_excel(excel_path, sheet_name='Weighted Variables', engine='openpyxl')

            if wgtd_df.empty:
                return True  # No weighted variables to import
        except:
            return True  # Sheet doesn't exist or is empty

        # Initialize weighted variables dictionary
        if not hasattr(model, 'wgtd_variables'):
            model.wgtd_variables = {}

        # Process the data
        current_var = None
        current_base = None

        for _, row in wgtd_df.iterrows():
            # Get the weighted variable name if present
            if not pd.isna(row['Weighted Variable']):
                current_var = row['Weighted Variable']
                current_base = row['Base Name'] if not pd.isna(row['Base Name']) else ''

                # Initialize the components dict for this variable
                if current_var not in model.wgtd_variables:
                    model.wgtd_variables[current_var] = {
                        'base_name': current_base,
                        'components': {}
                    }

            # Add the component if present
            if current_var and not pd.isna(row['Component Variable']) and not pd.isna(row['Coefficient']):
                component = row['Component Variable']
                coefficient = float(row['Coefficient'])

                model.wgtd_variables[current_var]['components'][component] = coefficient

        # Apply the weighted variables to the data
        from src.weighted_variables import create_weighted_variable_with_coefficients

        for var_name, var_info in model.wgtd_variables.items():
            # Skip if already in model data
            if var_name in model.model_data.columns:
                continue

            # Get base name and components
            base_name = var_info['base_name']
            components = var_info['components']

            # Create the weighted variable
            create_weighted_variable_with_coefficients(model, base_name, components)
            print(f"Imported weighted variable: {var_name}")

        return True
    except Exception as e:
        print(f"Error importing weighted variables: {str(e)}")
        return False