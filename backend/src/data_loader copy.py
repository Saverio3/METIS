import pandas as pd
import os
from pathlib import Path

class DataLoader:
    """
    A class to handle loading and basic processing of time series data from CSV or Excel files.
    Supports variable transformations (STA, SUB, MDV).
    """
    def __init__(self):
        self.data = None
        self.file_path = None
        self.transformations = {}  # Dictionary to store variable transformations
        
    def load_data(self, file_path):
        """
        Load data from a CSV or Excel file with dates in the first column.
        Looks for a 'Transformation' or 'Transformations' row at the top of the file.
        
        Parameters:
        -----------
        file_path : str
            Path to the data file (CSV or Excel)
            
        Returns:
        --------
        bool
            True if loading was successful, False otherwise
        """
        try:
            # Check if file exists
            if not os.path.exists(file_path):
                print(f"Error: File {file_path} not found.")
                return False
            
            # Determine file type
            file_extension = os.path.splitext(file_path)[1].lower()
            
            # Handle Excel files with transformation row specifically
            if file_extension in ['.xlsx', '.xls']:
                # Read the file using openpyxl directly to have more control
                import openpyxl
                
                # Load the workbook and first worksheet
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                
                # Get the transformation types from row 1
                transform_types = {}
                for col in range(2, ws.max_column + 1):  # Start from column B (index 2 in openpyxl)
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value is not None:
                        transform_types[col] = cell_value
                
                # Get the column names from row 2
                column_names = []
                for col in range(1, ws.max_column + 1):
                    column_names.append(ws.cell(row=2, column=col).value)
                
                # Map transformations to column names
                for col, trans in transform_types.items():
                    # Column names are 0-indexed in the list, but 1-indexed in openpyxl
                    col_name = column_names[col-1]
                    self.transformations[col_name] = trans
                
                # Now read the actual data with proper headers
                df = pd.read_excel(file_path, header=1)
                
                print(f"Loaded transformations for {len(self.transformations)} variables")
                
            elif file_extension == '.csv':
                # For CSV, try to handle transformation row
                # First read a few lines to check the structure
                with open(file_path, 'r') as f:
                    first_line = f.readline().strip()
                    second_line = f.readline().strip()
                
                # Check if first line might contain transformation info
                if 'transformation' in first_line.lower():
                    # Parse the first line to get transformations
                    first_cells = first_line.split(',')
                    second_cells = second_line.split(',')
                    
                    # Map transformations (skip first cell which is 'Transformation')
                    for i in range(1, min(len(first_cells), len(second_cells))):
                        if first_cells[i].strip():  # Not empty
                            self.transformations[second_cells[i].strip()] = first_cells[i].strip()
                    
                    # Now read the actual data with header in row 2
                    df = pd.read_csv(file_path, skiprows=1)
                else:
                    # Normal CSV without transformation row
                    df = pd.read_csv(file_path)
            else:
                print(f"Error: Unsupported file type '{file_extension}'. Only .csv, .xlsx, and .xls are supported.")
                return False
            
            # Check if 'Observation' column exists
            if 'Observation' not in df.columns:
                print("Error: File must contain an 'Observation' column with dates.")
                return False
            
            # Convert 'Observation' column to datetime
            try:
                df['Observation'] = pd.to_datetime(df['Observation'])
            except Exception as e:
                print(f"Error converting dates: {str(e)}")
                return False
            
            # Set 'Observation' as index
            df.set_index('Observation', inplace=True)
            
            # Store the data and file path
            self.data = df
            self.file_path = file_path
            
            return True
        
        except Exception as e:
            print(f"Error loading data: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
            
    def apply_transformation(self, variable_name, values, kpi_values=None):
        """
        Apply the specified transformation to a variable.
        
        Parameters:
        -----------
        variable_name : str
            Name of the variable to transform
        values : pandas.Series
            Values of the variable
        kpi_values : pandas.Series, optional
            Values of the KPI variable (needed for MDV transformation)
            
        Returns:
        --------
        pandas.Series
            Transformed values
        """
        if variable_name not in self.transformations:
            # No transformation, return original values
            return values
            
        trans_code = self.transformations[variable_name]
        
        if trans_code == 'STA':
            # Standardizing (STA): Dividing each observation by the mean of the region
            return values / values.mean() if values.mean() != 0 else values
        
        elif trans_code == 'SUB':
            # Subtracting (SUB): Subtracting from each observation the mean of the region
            return values - values.mean()
        
        elif trans_code == 'MDV':
            # Mean of Dependent Variable (MDV): Divides each variable by the mean of dependent variable
            if kpi_values is not None and len(kpi_values) > 0:
                kpi_mean = kpi_values.mean()
                return values / kpi_mean if kpi_mean != 0 else values
            else:
                print(f"Warning: Cannot apply MDV transformation for '{variable_name}' without KPI values")
                return values
            
        else:
            # Unknown transformation, return original values
            print(f"Warning: Unknown transformation code '{trans_code}' for variable '{variable_name}'")
            return values
            
    def get_transformation(self, variable_name):
        """
        Get the transformation code for a variable.
        
        Parameters:
        -----------
        variable_name : str
            Name of the variable
            
        Returns:
        --------
        str or None
            Transformation code or None if no transformation is defined
        """
        return self.transformations.get(variable_name, None)
    
    def set_transformation(self, variable_name, transformation):
      """
    Set the transformation code for a variable.
    
    Parameters:
    -----------
    variable_name : str
        Name of the variable
    transformation : str
        Transformation code to apply (STA, SUB, MDV)
    """
    if transformation == 'NONE':
        # Remove transformation if exists
        if variable_name in self.transformations:
            del self.transformations[variable_name]
    else:
        # Set or update transformation
        self.transformations[variable_name] = transformation
    
        return True
        
    def load_csv(self, file_path):
        """
        Backward compatibility method for loading data.
        Redirects to load_data method.
        """
        return self.load_data(file_path)
    
    def register_transformation(self, variable_name, transformation_info):
      """
      Register detailed transformation information for a variable.
      This is used for exporting and recreating transformations.
    
      Parameters:
      -----------
      variable_name : str
          Name of the transformed variable
      transformation_info : dict
          Dictionary with transformation details
        
      Returns:
      --------
      bool
          True if successful
      """
    # Initialize transformations_registry if it doesn't exist
    if not hasattr(self, 'transformations_registry'):
        self.transformations_registry = {}
        
    # Store the transformation info
    self.transformations_registry[variable_name] = transformation_info
    
    return True
    
    def get_data(self):
        """
        Get the loaded data.
        
        Returns:
        --------
        pandas.DataFrame or None
            The loaded data or None if no data has been loaded
        """
        return self.data
    
    def get_variable_names(self):
        """
        Get the names of variables in the dataset.
        
        Returns:
        --------
        list or None
            List of column names or None if no data has been loaded
        """
        if self.data is not None:
            return list(self.data.columns)
        return None
    
    def get_transformations(self):
        """
        Get the transformations dictionary.
        
        Returns:
        --------
        dict
            Dictionary mapping variable names to transformation codes
        """
        return self.transformations
    
    def get_summary(self):
        """
        Get a summary of the loaded data.
        
        Returns:
        --------
        dict or None
            Dictionary with summary information or None if no data has been loaded
        """
        if self.data is None:
            return None
        
        return {
            'file_name': os.path.basename(self.file_path) if self.file_path else 'Unknown',
            'n_observations': len(self.data),
            'date_range': f"{self.data.index.min().strftime('%Y-%m-%d')} to {self.data.index.max().strftime('%Y-%m-%d')}",
            'variables': len(self.data.columns),
            'variable_names': list(self.data.columns),
            'transformations': len(self.transformations)
        }
        
    def filter_date_range(self, start_date=None, end_date=None):
        """
        Filter data to include only observations within the specified date range.
        
        Parameters:
        -----------
        start_date : str or datetime, optional
            The start date of the range (inclusive)
        end_date : str or datetime, optional
            The end date of the range (inclusive)
            
        Returns:
        --------
        pandas.DataFrame
            Filtered dataframe
        """
        if self.data is None:
            print("No data loaded. Please load data first.")
            return None
            
        filtered_data = self.data.copy()
        
        if start_date is not None:
            try:
                start_date = pd.to_datetime(start_date)
                filtered_data = filtered_data[filtered_data.index >= start_date]
            except Exception as e:
                print(f"Error filtering start date: {str(e)}")
                
        if end_date is not None:
            try:
                end_date = pd.to_datetime(end_date)
                filtered_data = filtered_data[filtered_data.index <= end_date]
            except Exception as e:
                print(f"Error filtering end date: {str(e)}")
                
        return filtered_data
    
    def save_metadata(self, file_path=None):
        """
        Save data metadata to a file for reference.
        
        Parameters:
        -----------
        file_path : str, optional
            Path where to save the metadata. If None, uses default location.
            
        Returns:
        --------
        str
            Path to the saved metadata file
        """
        if self.data is None:
            print("No data to save metadata for.")
            return None
            
        try:
            # If file_path is not provided, create one
            if file_path is None:
                Path('data').mkdir(parents=True, exist_ok=True)
                file_name = os.path.splitext(os.path.basename(self.file_path))[0] if self.file_path else 'data'
                file_path = os.path.join('data', f"{file_name}_metadata.txt")
            
            with open(file_path, 'w') as f:
                f.write(f"Data File: {self.file_path}\n")
                f.write(f"Variables: {', '.join(self.data.columns)}\n")
                f.write(f"Observations: {len(self.data)}\n")
                f.write(f"Date Range: {self.data.index.min().strftime('%Y-%m-%d')} to {self.data.index.max().strftime('%Y-%m-%d')}\n")
                
                # Add transformation information
                f.write(f"\nTransformations: {len(self.transformations)} variables\n")
                if self.transformations:
                    f.write("Variable Transformations:\n")
                    for var, trans in self.transformations.items():
                        f.write(f"  {var}: {trans}\n")
                
                # Add basic statistics for each variable
                f.write("\nBasic Statistics:\n")
                f.write("================\n")
                for col in self.data.columns:
                    f.write(f"\n{col}:\n")
                    f.write(f"  Mean: {self.data[col].mean():.4f}\n")
                    f.write(f"  Std Dev: {self.data[col].std():.4f}\n")
                    f.write(f"  Min: {self.data[col].min():.4f}\n")
                    f.write(f"  Max: {self.data[col].max():.4f}\n")
                    f.write(f"  Missing Values: {self.data[col].isna().sum()}\n")
                    if col in self.transformations:
                        f.write(f"  Transformation: {self.transformations[col]}\n")
            
            print(f"Metadata saved to {file_path}")
            return file_path
            
        except Exception as e:
            print(f"Error saving metadata: {str(e)}")
            return None